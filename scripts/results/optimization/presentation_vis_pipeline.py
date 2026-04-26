from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import pandas as pd
import yaml


def _repo_root_from(path: Path) -> Path:
    for candidate in (path, *path.parents):
        if (candidate / ".git").exists():
            return candidate
    return path.parents[2]


ROOT = _repo_root_from(Path(__file__).resolve())
BASE_OPT_CONFIG = ROOT / "configs/scheduling/base_optimization.yaml"
DEFAULT_SCREENING_SOURCE = ROOT / "results/thesis_optimization_results/results/mtlsh_topk_screening/global/screening_sets.json"

FORMULATION_SPECS = {
    "A": {
        "id": "ed",
        "name": "ED only",
        "description": "Economic dispatch only.",
        "uses_line": False,
        "uses_n1": False,
        "uses_surrogate": False,
    },
    "B": {
        "id": "ed_line",
        "name": "ED + line constraints",
        "description": "Economic dispatch + base-case PTDF line constraints.",
        "uses_line": True,
        "uses_n1": False,
        "uses_surrogate": False,
    },
    "C": {
        "id": "ed_line_n1",
        "name": "ED + line + preventive N-1",
        "description": "ED + line + preventive N-1 security.",
        "uses_line": True,
        "uses_n1": True,
        "uses_surrogate": False,
    },
    "D": {
        "id": "ed_surrogate",
        "name": "ED + surrogate constraints",
        "description": "ED + surrogate dynamic-security constraints.",
        "uses_line": False,
        "uses_n1": False,
        "uses_surrogate": True,
    },
    "E": {
        "id": "ed_line_n1_surrogate",
        "name": "Full formulation",
        "description": "ED + line + preventive N-1 + surrogate constraints.",
        "uses_line": True,
        "uses_n1": True,
        "uses_surrogate": True,
    },
}

FORMULATION_CODE_BY_ID = {
    "ed": "A",
    "ed_line": "B",
    "ed_line_n1": "C",
    "ed_surrogate": "D",
    "ed_line_n1_surrogate": "E",
    "ed_line_n1_surrogate_redispatch": "E",
}

FORMULATION_LABEL_BY_ID = {
    "ed": "A: ED",
    "ed_line": "B: ED + Line",
    "ed_line_n1": "C: ED + Line + N-1",
    "ed_surrogate": "D: ED + Surrogate",
    "ed_line_n1_surrogate": "E: Full preventive",
    "ed_line_n1_surrogate_redispatch": "E: Full preventive + Redispatch",
}

DEFAULT_CONFIG: dict[str, Any] = {
    "run": {
        "case_label": "vis_case",
        "output_root": "results/presentation_vis",
        "random_seed": 42,
        "save_csv": True,
        "save_metadata": True,
        "run_replay": True,
        "generated_configs_root": "",
    },
    "system": {
        "sbase": 100.0,
        "base_load_scale": 0.75,
        "load_step_time_s": 1.0,
        "disturbance_family": "global_mismatch",
        "disturbance_scale": 1.2,
        "zone_id": None,
        "contingency_label": "none",
        "contingency_mode": "none",
    },
    "formulations": {
        "run_A_ed_only": True,
        "run_B_ed_line": True,
        "run_C_ed_line_n1": True,
        "run_D_ed_surrogate": True,
        "run_E_full": True,
        "run_no_vi_reference": False,
    },
    "security_limits": {
        "rocof_lim_hz_per_s": 1.0,
        "delta_f_lim_hz": 0.8,
        "enforce_delta_p_limits": True,
        "delta_p_mode": "existing_repo_mode",
    },
    "network_security": {
        "enforce_line_limits": True,
        "enforce_n1_redispatch": False,
        "topk_contingencies": None,
        "screened_contingency_set_name": None,
        "screened_contingency_source": str(DEFAULT_SCREENING_SOURCE),
    },
    "surrogate": {
        "enabled": True,
        "model_bundle_path": "",
        "scaler_bundle_path": "",
        "export_name": "retained_surrogate",
        "embedding_mode": "milp",
        "retained_architecture_name": "",
    },
    "vis_bounds": {
        "m_min": 0.0,
        "m_max": 8.0,
        "d_min": 0.0,
        "d_max": 6.0,
    },
    "comparison": {
        "run_default_bounds_case": True,
        "run_tight_bounds_case": False,
        "tight_case_suffix": "tight",
        "default_case_suffix": "default",
    },
    "logging": {
        "save_solver_logs": True,
        "save_intermediate_schedule_tables": True,
        "save_time_series": True,
        "trajectory_time_window_s": [0.0, 20.0],
        "trajectory_sampling_step": 0.05,
        "save_per_unit_and_mw_versions": True,
        "log_tail_lines": 120,
    },
}


class PipelineError(RuntimeError):
    pass


def _deep_merge(base: dict[str, Any], patch: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = dict(base)
    for key, value in patch.items():
        if isinstance(value, dict) and isinstance(out.get(key), dict):
            out[key] = _deep_merge(dict(out[key]), value)  # type: ignore[arg-type]
        else:
            out[key] = value
    return out


def _load_yaml(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _write_yaml(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        yaml.safe_dump(payload, f, sort_keys=False)


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)


def _write_df(path: Path, df: pd.DataFrame, columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    out = df.copy()
    for col in columns:
        if col not in out.columns:
            out[col] = np.nan
    out = out.loc[:, columns]
    out.to_csv(path, index=False)


def _resolve(path_like: str | Path, base: Path = ROOT) -> Path:
    p = Path(path_like)
    return p if p.is_absolute() else (base / p).resolve()


def _safe_float(value: Any, default: float = np.nan) -> float:
    try:
        if value is None:
            return float(default)
        return float(value)
    except Exception:
        return float(default)


def _parse_line_uid(label: Any) -> int:
    if label is None:
        raise PipelineError("contingency_label is required for line-based settings.")
    text = str(label).strip()
    if text.isdigit():
        return int(text)
    match = re.search(r"(\d+)", text)
    if not match:
        raise PipelineError(f"Could not parse a line uid from contingency_label='{text}'.")
    return int(match.group(1))


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _git_commit_hash() -> str:
    try:
        out = subprocess.check_output(["git", "rev-parse", "HEAD"], cwd=str(ROOT), text=True)
        return out.strip()
    except Exception:
        return ""


def _normalize_metric_name(name: Any) -> str:
    raw = str(name)
    mapping = {
        "rocof_COI": "rocof_coi",
        "dev_COI": "delta_f_coi",
        "Delta_P_IBR_1": "delta_p_ibr_1",
        "Delta_P_IBR_2": "delta_p_ibr_2",
        "Delta_P_IBR_3": "delta_p_ibr_3",
        "Delta_P_IBR_4": "delta_p_ibr_4",
    }
    return mapping.get(raw, raw.lower())


def _apply_defaults(cfg: dict[str, Any]) -> dict[str, Any]:
    return _deep_merge(DEFAULT_CONFIG, cfg)


def _validate_config(cfg: dict[str, Any]) -> None:
    required_sections = [
        "run",
        "system",
        "formulations",
        "security_limits",
        "network_security",
        "surrogate",
        "vis_bounds",
        "comparison",
        "logging",
    ]
    missing = [key for key in required_sections if key not in cfg]
    if missing:
        raise PipelineError(f"Missing required config sections: {missing}")

    run_cfg = dict(cfg.get("run", {}) or {})
    if not str(run_cfg.get("case_label", "")).strip():
        raise PipelineError("run.case_label must be non-empty.")

    comparison_cfg = dict(cfg.get("comparison", {}) or {})
    if not bool(comparison_cfg.get("run_default_bounds_case", False)) and not bool(
        comparison_cfg.get("run_tight_bounds_case", False)
    ):
        raise PipelineError(
            "At least one of comparison.run_default_bounds_case or comparison.run_tight_bounds_case must be true."
        )

    system_cfg = dict(cfg.get("system", {}) or {})
    family = str(system_cfg.get("disturbance_family", "")).strip().lower()
    if family not in {"global_mismatch", "zone_mismatch", "line_outage", "mixed"}:
        raise PipelineError(
            "system.disturbance_family must be one of: global_mismatch, zone_mismatch, line_outage, mixed."
        )
    if family == "zone_mismatch" and system_cfg.get("zone_id") in {None, "", "null"}:
        raise PipelineError("system.zone_id is required for disturbance_family=zone_mismatch.")

    cont_mode = str(system_cfg.get("contingency_mode", "none")).strip().lower()
    valid_modes = {"none", "single_line", "screened_set", "topk"}
    if cont_mode not in valid_modes:
        raise PipelineError(f"system.contingency_mode must be one of {sorted(valid_modes)}")
    if cont_mode == "single_line" and str(system_cfg.get("contingency_label", "")).strip().lower() in {"", "none"}:
        raise PipelineError("system.contingency_label is required for contingency_mode=single_line.")

    if cont_mode == "topk":
        topk = cfg.get("network_security", {}).get("topk_contingencies")
        if topk is None or int(topk) <= 0:
            raise PipelineError("network_security.topk_contingencies must be a positive integer for contingency_mode=topk.")

    if cont_mode == "screened_set":
        name = cfg.get("network_security", {}).get("screened_contingency_set_name")
        if not str(name or "").strip():
            raise PipelineError(
                "network_security.screened_contingency_set_name is required for contingency_mode=screened_set."
            )


def _load_screening_sets(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise PipelineError(
            f"Screened contingency source not found: {path}. "
            "Run results/thesis_optimization_results/scripts/09_run_mtlsh_topk_screening.sh first or update network_security.screened_contingency_source."
        )
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _resolve_n1_include_uids(cfg: dict[str, Any]) -> list[int] | None:
    system_cfg = dict(cfg.get("system", {}) or {})
    mode = str(system_cfg.get("contingency_mode", "none")).strip().lower()
    if mode == "none":
        return None

    if mode == "single_line":
        return [_parse_line_uid(system_cfg.get("contingency_label"))]

    net_cfg = dict(cfg.get("network_security", {}) or {})
    source = _resolve(str(net_cfg.get("screened_contingency_source", DEFAULT_SCREENING_SOURCE)))
    payload = _load_screening_sets(source)

    if mode == "topk":
        topk = int(net_cfg.get("topk_contingencies"))
        rows = list(payload.get("available_screened_outages") or [])
        rows = sorted(rows, key=lambda r: int(r.get("selection_rank", 10**9)))
        return [int(row["line_uid"]) for row in rows[:topk]]

    if mode == "screened_set":
        target = str(net_cfg.get("screened_contingency_set_name", "")).strip()
        generated = list(payload.get("generated_runs") or [])
        for row in generated:
            formulation_id = str(row.get("formulation_id", "")).strip()
            screening_label = str(row.get("screening_label", "")).strip()
            if target in {formulation_id, screening_label}:
                return [int(v) for v in list(row.get("line_uids") or [])]

        # Accept explicit comma-separated list as fallback.
        tokens = [token.strip() for token in target.split(",") if token.strip()]
        if tokens and all(re.search(r"\d+", token) for token in tokens):
            return [_parse_line_uid(token) for token in tokens]

        raise PipelineError(
            f"Could not resolve screened contingency set '{target}' from {source}."
        )

    return None


def _build_scenario_overrides(cfg: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any]]:
    system_cfg = dict(cfg.get("system", {}) or {})
    family = str(system_cfg.get("disturbance_family", "global_mismatch")).strip().lower()

    scenario = {
        "base_scale": float(system_cfg.get("base_load_scale", 0.75)),
        "step_scale": float(system_cfg.get("disturbance_scale", 1.2)),
        "load_step_time": float(system_cfg.get("load_step_time_s", 1.0)),
    }
    features: dict[str, Any] = {"contingency_mode": "load_mismatch"}

    if family == "zone_mismatch":
        scenario["load_step_target_owners"] = [str(system_cfg.get("zone_id"))]
    elif family in {"line_outage", "mixed"}:
        line_uid = _parse_line_uid(system_cfg.get("contingency_label"))
        scenario["contingency_line_uid"] = int(line_uid)
        features["contingency_mode"] = "line"
        features["contingency_line_uid"] = int(line_uid)

    return scenario, features


def _build_scaler_overrides(path_like: str) -> dict[str, Any]:
    path = _resolve(path_like)
    if path.is_dir():
        return {
            "scalers": {
                "x_scaler_path": str(path / "x_scaler.pkl"),
                "y_scaler_path": str(path / "y_scaler.pkl"),
            }
        }

    name = path.name.lower()
    if "x_scaler" in name:
        return {
            "scalers": {
                "x_scaler_path": str(path),
                "y_scaler_path": str(path.parent / "y_scaler.pkl"),
            }
        }
    if "y_scaler" in name:
        return {
            "scalers": {
                "x_scaler_path": str(path.parent / "x_scaler.pkl"),
                "y_scaler_path": str(path),
            }
        }

    # Fallback: assume a bundle directory next to the provided path.
    return {
        "scalers": {
            "x_scaler_path": str(path.parent / "x_scaler.pkl"),
            "y_scaler_path": str(path.parent / "y_scaler.pkl"),
        }
    }


def _build_formulation_runs(
    cfg: dict[str, Any],
    *,
    bounds_case: str,
    apply_tight_limits: bool,
    n1_include_uids: list[int] | None,
) -> tuple[list[dict[str, Any]], dict[str, str], list[str]]:
    formulations_cfg = dict(cfg.get("formulations", {}) or {})
    net_cfg = dict(cfg.get("network_security", {}) or {})
    surrogate_cfg = dict(cfg.get("surrogate", {}) or {})
    security_cfg = dict(cfg.get("security_limits", {}) or {})
    vis_cfg = dict(cfg.get("vis_bounds", {}) or {})

    requested_codes = []
    if bool(formulations_cfg.get("run_A_ed_only", False)):
        requested_codes.append("A")
    if bool(formulations_cfg.get("run_B_ed_line", False)):
        requested_codes.append("B")
    if bool(formulations_cfg.get("run_C_ed_line_n1", False)):
        requested_codes.append("C")
    if bool(formulations_cfg.get("run_D_ed_surrogate", False)):
        requested_codes.append("D")
    if bool(formulations_cfg.get("run_E_full", False)):
        requested_codes.append("E")

    if not requested_codes:
        raise PipelineError("No formulations selected in config.formulations.")

    runs: list[dict[str, Any]] = []
    map_id_to_code: dict[str, str] = {}
    notes: list[str] = []

    base_cfg = _load_yaml(BASE_OPT_CONFIG)
    base_y_min = list(base_cfg.get("bounds", {}).get("y_min", [-1.0, -0.8, -5.0, -5.0, -5.0, -5.0]))
    base_y_max = list(base_cfg.get("bounds", {}).get("y_max", [1.0, 0.8, 5.0, 5.0, 5.0, 5.0]))

    y_min = [float(v) for v in base_y_min]
    y_max = [float(v) for v in base_y_max]
    if apply_tight_limits:
        rocof_lim = abs(float(security_cfg.get("rocof_lim_hz_per_s", 1.0)))
        delta_f_lim = abs(float(security_cfg.get("delta_f_lim_hz", 0.8)))
        y_min[0], y_max[0] = -rocof_lim, rocof_lim
        y_min[1], y_max[1] = -delta_f_lim, delta_f_lim

    if not bool(security_cfg.get("enforce_delta_p_limits", True)):
        for i in range(2, min(len(y_min), len(y_max))):
            y_min[i] = -1.0e6
            y_max[i] = 1.0e6

    for code in requested_codes:
        spec = FORMULATION_SPECS[code]
        formulation_id = str(spec["id"])

        uses_surrogate = bool(spec["uses_surrogate"])
        if uses_surrogate and not bool(surrogate_cfg.get("enabled", True)):
            notes.append(f"Skipped formulation {code} because surrogate.enabled=false")
            continue

        use_line = bool(spec["uses_line"]) and bool(net_cfg.get("enforce_line_limits", True))
        use_n1 = bool(spec["uses_n1"]) and use_line
        use_redispatch = bool(use_n1 and net_cfg.get("enforce_n1_redispatch", False))

        constraints = {
            "use_input": bool(uses_surrogate),
            "use_output": bool(uses_surrogate),
            "use_nn": bool(uses_surrogate),
            "use_line": bool(use_line),
            "use_n1": bool(use_n1),
            "use_n1_redispatch": bool(use_redispatch),
            "use_ed": True,
            "nn_mode": str(surrogate_cfg.get("embedding_mode", "milp")),
            "enforce_dispatch_output_link": True,
        }

        if n1_include_uids and use_n1:
            constraints["include_n1_line_uids"] = [int(v) for v in n1_include_uids]

        overrides: dict[str, Any] = {
            "constraints": constraints,
            "bounds": {
                "M_bounds": [float(vis_cfg.get("m_min", 0.0)), float(vis_cfg.get("m_max", 8.0))],
                "D_bounds": [float(vis_cfg.get("d_min", 0.0)), float(vis_cfg.get("d_max", 6.0))],
                "y_min": y_min,
                "y_max": y_max,
            },
            "formulation": {
                "id": formulation_id,
                "name": str(spec["name"]),
                "description": str(spec["description"]),
            },
        }

        model_bundle = str(surrogate_cfg.get("model_bundle_path", "")).strip()
        if model_bundle:
            overrides.setdefault("model", {})
            overrides["model"]["model_dir"] = str(_resolve(model_bundle))

        arch_name = str(surrogate_cfg.get("retained_architecture_name", "")).strip()
        if arch_name:
            overrides.setdefault("model", {})
            overrides["model"]["type"] = arch_name

        scaler_bundle = str(surrogate_cfg.get("scaler_bundle_path", "")).strip()
        if scaler_bundle:
            overrides = _deep_merge(overrides, _build_scaler_overrides(scaler_bundle))

        runs.append(
            {
                "id": formulation_id,
                "name": f"{code}: {spec['name']}",
                "description": f"Presentation VIS run ({bounds_case})",
                "base_config": str(BASE_OPT_CONFIG),
                "overrides": overrides,
            }
        )
        map_id_to_code[formulation_id] = code

    if not runs:
        raise PipelineError("All requested formulations were skipped. Check surrogate/formulation flags.")

    return runs, map_id_to_code, notes


def _resolve_replay_pq_targets(case_path: Path, zone_id: Any) -> list[str]:
    if zone_id in {None, "", "null"}:
        return []
    if not case_path.exists() or case_path.suffix.lower() != ".xlsx":
        return []

    owner = str(zone_id).strip()
    wb = openpyxl.load_workbook(case_path, data_only=True, read_only=True)
    if "PQ" not in wb.sheetnames:
        return []

    ws = wb["PQ"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(v) for v in rows[0]]
    idx_map = {name: i for i, name in enumerate(headers)}
    if "name" not in idx_map or "owner" not in idx_map:
        return []

    enabled_idx = idx_map.get("u")
    out: list[str] = []
    for row in rows[1:]:
        if not row:
            continue
        if enabled_idx is not None:
            enabled = row[enabled_idx]
            if enabled is not None and int(enabled) != 1:
                continue
        row_owner = str(row[idx_map["owner"]]).strip()
        if row_owner == owner:
            out.append(str(row[idx_map["name"]]).strip())

    return sorted(set(v for v in out if v))


def _build_replay_contingency(cfg: dict[str, Any], base_opt: dict[str, Any]) -> dict[str, Any]:
    system_cfg = dict(cfg.get("system", {}) or {})
    family = str(system_cfg.get("disturbance_family", "global_mismatch")).strip().lower()
    cont_mode = str(system_cfg.get("contingency_mode", "none")).strip().lower()

    if family in {"line_outage", "mixed"} or cont_mode == "single_line":
        return {
            "type": "line_trip",
            "time": float(system_cfg.get("load_step_time_s", 1.0)),
            "line_uid": int(_parse_line_uid(system_cfg.get("contingency_label"))),
        }

    if cont_mode == "none" and family == "line_outage":
        return {"type": "none"}

    contingency = {
        "type": "load_step",
        "time": float(system_cfg.get("load_step_time_s", 1.0)),
        "scale": float(system_cfg.get("disturbance_scale", base_opt.get("scenario", {}).get("step_scale", 1.2))),
    }

    if family == "zone_mismatch":
        case_path = _resolve(str(base_opt.get("system", {}).get("case", "")))
        targets = _resolve_replay_pq_targets(case_path, zone_id=system_cfg.get("zone_id"))
        if targets:
            contingency["pq_targets"] = targets
    return contingency


def _build_suite_config(
    cfg: dict[str, Any],
    *,
    bounds_case: str,
    out_root: Path,
    apply_tight_limits: bool,
) -> tuple[dict[str, Any], dict[str, str], list[str], dict[str, Any]]:
    scenario_overrides, feature_overrides = _build_scenario_overrides(cfg)
    n1_include_uids = _resolve_n1_include_uids(cfg)
    runs, map_id_to_code, notes = _build_formulation_runs(
        cfg,
        bounds_case=bounds_case,
        apply_tight_limits=apply_tight_limits,
        n1_include_uids=n1_include_uids,
    )

    case_label = str(cfg["run"]["case_label"])
    raw_root = out_root / "raw"
    suite_cfg = {
        "name": f"presentation_vis__{case_label}__{bounds_case}",
        "baseline_id": runs[0]["id"],
        "results_root": str(raw_root / "formulations"),
        "output": {
            "summary_csv": str(raw_root / "suite_summary.csv"),
            "summary_markdown": str(raw_root / "suite_summary.md"),
            "summary_json": str(raw_root / "suite_summary.json"),
        },
        "scenarios": [
            {
                "id": case_label,
                "name": case_label,
                "description": "Presentation VIS scenario",
                "overrides": {
                    "scenario": scenario_overrides,
                    "features": feature_overrides,
                },
            }
        ],
        "runs": runs,
    }

    base_opt = _load_yaml(BASE_OPT_CONFIG)
    replay_contingency = _build_replay_contingency(cfg, base_opt)
    return suite_cfg, map_id_to_code, notes, replay_contingency


def _run_command(cmd: list[str], *, cwd: Path, env: dict[str, str] | None = None) -> None:
    print("[presentation_vis] $", " ".join(cmd))
    subprocess.run(cmd, cwd=str(cwd), env=env, check=True)


def _load_suite_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise PipelineError(f"Suite summary JSON was not produced: {path}")
    with path.open("r", encoding="utf-8") as f:
        payload = json.load(f)
    return list(payload.get("rows") or [])


def _load_summary_payload(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _load_unit_catalog() -> dict[int, dict[str, Any]]:
    cost_path = ROOT / "configs/shared/ieee39_regcv1_dispatch_costs.yaml"
    if not cost_path.exists():
        return {}

    payload = _load_yaml(cost_path)
    out: dict[int, dict[str, Any]] = {}
    ref = dict(payload.get("dispatch_order_reference", {}) or {})
    labels = list(ref.get("labels") or [])
    buses = list(ref.get("buses") or [])
    row_map = {int(row["bus"]): row for row in list(payload.get("generators") or []) if row.get("bus") is not None}

    if labels and buses and len(labels) == len(buses):
        for idx, (label, bus) in enumerate(zip(labels, buses)):
            row = row_map.get(int(bus), {})
            out[int(idx)] = {
                "unit_name": str(label),
                "unit_type": str(row.get("type", "unknown")),
                "bus": int(bus),
            }

    if out:
        return out

    generators = list(payload.get("generators") or [])
    for idx, row in enumerate(generators):
        out[int(idx)] = {
            "unit_name": str(row.get("label", f"unit_{idx + 1}")),
            "unit_type": str(row.get("type", "unknown")),
            "bus": int(row.get("bus", idx + 1)),
        }
    return out


def _collect_optimization_tables(
    *,
    rows: list[dict[str, Any]],
    map_id_to_code: dict[str, str],
    case_label: str,
    bounds_case: str,
    sbase: float,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, dict[int, dict[str, Any]]], list[str]]:
    formulations: list[dict[str, Any]] = []
    costs: list[dict[str, Any]] = []
    setpoints: list[dict[str, Any]] = []
    errors: list[str] = []
    per_formulation_ibr_map: dict[str, dict[int, dict[str, Any]]] = {}

    unit_catalog = _load_unit_catalog()

    for row in rows:
        formulation_id = str(row.get("formulation_id", "")).strip()
        formulation_code = map_id_to_code.get(formulation_id, FORMULATION_CODE_BY_ID.get(formulation_id, formulation_id))
        summary_json_raw = str(row.get("summary_json", "")).strip()
        status = str(row.get("status", "")).strip()

        summary_payload: dict[str, Any] = {}
        if summary_json_raw:
            summary_path = Path(summary_json_raw)
            if summary_path.exists():
                summary_payload = _load_summary_payload(summary_path)

        switches = dict(summary_payload.get("constraint_switches", {}) or {})
        solver_stats = dict(summary_payload.get("solver_stats", {}) or {})

        solve_time_ms = _safe_float(solver_stats.get("solve_time_sec"), np.nan)
        solve_time_ms = solve_time_ms * 1e3 if np.isfinite(solve_time_ms) else np.nan

        formulations.append(
            {
                "formulation_code": formulation_code,
                "formulation_name": str(summary_payload.get("formulation_name", row.get("formulation_name", formulation_id))),
                "includes_line_constraints": int(bool(switches.get("use_line", row.get("use_line", False)))),
                "includes_n1_redispatch": int(bool(switches.get("use_n1_redispatch", row.get("use_n1_redispatch", False)))),
                "includes_surrogate": int(bool(switches.get("use_nn", row.get("use_nn", False)))),
                "includes_vis": int(bool(switches.get("use_input", False))),
                "is_feasible": int(str(status).startswith("optimal")),
                "solve_time_ms": solve_time_ms,
                "case_label": case_label,
                "bounds_case": bounds_case,
            }
        )

        dispatch_only = _safe_float(summary_payload.get("objective_dispatch_only"), np.nan)
        reserve_only = _safe_float(summary_payload.get("objective_reserve_only"), np.nan)
        reserve_postcont = _safe_float(summary_payload.get("objective_reserve_postcont_only"), np.nan)
        if np.isfinite(reserve_only) and np.isfinite(reserve_postcont):
            reserve_only = float(reserve_only - reserve_postcont)
        tie_breaker = _safe_float(summary_payload.get("objective_tie_breaker"), 0.0)
        md_reg = _safe_float(summary_payload.get("objective_md_regularization"), 0.0)

        objective_total = np.nan
        if np.isfinite(dispatch_only) and np.isfinite(reserve_only):
            objective_total = float(dispatch_only + reserve_only)
        elif np.isfinite(dispatch_only):
            objective_total = float(dispatch_only)
        else:
            objective_total = _safe_float(summary_payload.get("objective"), np.nan)

        costs.append(
            {
                "formulation_code": formulation_code,
                "objective_cost_kusd": objective_total,
                "dispatch_only_cost_kusd": dispatch_only,
                "reserve_cost_kusd": reserve_only,
                "other_cost_component": float(tie_breaker + md_reg),
                "solve_time_ms": solve_time_ms,
                "case_label": case_label,
                "bounds_case": bounds_case,
            }
        )

        artifacts = dict(summary_payload.get("artifacts", {}) or {})
        dispatch_path_raw = str(artifacts.get("dispatch_impact_csv", row.get("dispatch_impact_csv", ""))).strip()
        if not dispatch_path_raw:
            continue
        dispatch_path = Path(dispatch_path_raw)
        if not dispatch_path.exists():
            errors.append(f"Missing dispatch_impact_csv for {formulation_id}: {dispatch_path}")
            continue

        dispatch_df = pd.read_csv(dispatch_path)
        dispatch_df.columns = [str(col).strip() for col in dispatch_df.columns]

        gen_df = dispatch_df.loc[dispatch_df.get("row_type", "").astype(str) == "generator_dispatch"].copy()
        ibr_df = dispatch_df.loc[dispatch_df.get("row_type", "").astype(str) == "ibr_summary"].copy()

        md_by_gen: dict[int, tuple[float, float]] = {}
        local_to_gen: dict[int, int] = {}
        for _, irow in ibr_df.iterrows():
            gen_idx = int(_safe_float(irow.get("gen_index"), np.nan)) if np.isfinite(_safe_float(irow.get("gen_index"), np.nan)) else None
            local_idx = int(_safe_float(irow.get("index"), np.nan)) if np.isfinite(_safe_float(irow.get("index"), np.nan)) else None
            m_opt = _safe_float(irow.get("M_opt"), np.nan)
            d_opt = _safe_float(irow.get("D_opt"), np.nan)
            if gen_idx is not None:
                md_by_gen[int(gen_idx)] = (m_opt, d_opt)
            if local_idx is not None and gen_idx is not None:
                local_to_gen[int(local_idx)] = int(gen_idx)

        per_formulation_ibr_map[formulation_code] = {}
        for local_idx, gen_idx in local_to_gen.items():
            meta = unit_catalog.get(gen_idx, {"unit_name": f"unit_{gen_idx + 1}", "unit_type": "unknown"})
            scheduled_pu = np.nan
            row_match = gen_df.loc[pd.to_numeric(gen_df.get("index"), errors="coerce") == float(gen_idx)]
            if not row_match.empty:
                scheduled_pu = _safe_float(row_match.iloc[0].get("pg_opt"), np.nan)
            per_formulation_ibr_map[formulation_code][local_idx + 1] = {
                "unit_name": str(meta.get("unit_name", f"unit_{gen_idx + 1}")),
                "unit_type": str(meta.get("unit_type", "unknown")),
                "scheduled_dispatch_pu": scheduled_pu,
            }

        for _, grow in gen_df.iterrows():
            idx = int(_safe_float(grow.get("index"), np.nan)) if np.isfinite(_safe_float(grow.get("index"), np.nan)) else None
            if idx is None:
                continue
            meta = unit_catalog.get(int(idx), {"unit_name": f"unit_{idx + 1}", "unit_type": "unknown"})
            pg_opt = _safe_float(grow.get("pg_opt"), np.nan)
            reserve_up = _safe_float(grow.get("reserve_up"), np.nan)
            pg_min = _safe_float(grow.get("pg_min"), np.nan)
            reserve_down = pg_opt - pg_min if np.isfinite(pg_opt) and np.isfinite(pg_min) else np.nan

            m_opt, d_opt = md_by_gen.get(int(idx), (np.nan, np.nan))

            setpoints.append(
                {
                    "formulation_code": formulation_code,
                    "unit_name": str(meta.get("unit_name", f"unit_{idx + 1}")),
                    "unit_type": str(meta.get("unit_type", "unknown")),
                    "scheduled_dispatch_mw": pg_opt * sbase if np.isfinite(pg_opt) else np.nan,
                    "upward_headroom_mw": reserve_up * sbase if np.isfinite(reserve_up) else np.nan,
                    "downward_headroom_mw": reserve_down * sbase if np.isfinite(reserve_down) else np.nan,
                    "virtual_inertia": m_opt,
                    "virtual_damping": d_opt,
                    "reserve_up_mw": reserve_up * sbase if np.isfinite(reserve_up) else np.nan,
                    "reserve_down_mw": reserve_down * sbase if np.isfinite(reserve_down) else np.nan,
                    "case_label": case_label,
                    "bounds_case": bounds_case,
                }
            )

    formulations_df = pd.DataFrame(formulations)
    costs_df = pd.DataFrame(costs)
    setpoints_df = pd.DataFrame(setpoints)
    return formulations_df, costs_df, setpoints_df, per_formulation_ibr_map, errors


def _collect_replay_metrics(
    *,
    replay_detail_csv: Path,
    map_id_to_code: dict[str, str],
    cfg: dict[str, Any],
    case_label: str,
    bounds_case: str,
) -> pd.DataFrame:
    columns = [
        "formulation_code",
        "metric_name",
        "predicted_value",
        "replayed_value",
        "limit_value",
        "is_within_limit",
        "case_label",
        "bounds_case",
        "contingency_label",
    ]
    if not replay_detail_csv.exists():
        return pd.DataFrame(columns=columns)

    df = pd.read_csv(replay_detail_csv)
    if df.empty:
        return pd.DataFrame(columns=columns)

    sec_cfg = dict(cfg.get("security_limits", {}) or {})
    delta_p_mode = str(sec_cfg.get("delta_p_mode", "existing_repo_mode")).strip().lower()
    contingency_label = str(cfg.get("system", {}).get("contingency_label", "none"))

    out_rows: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        formulation_id = str(row.get("formulation_id", "")).strip()
        formulation_code = map_id_to_code.get(formulation_id, FORMULATION_CODE_BY_ID.get(formulation_id, formulation_id))
        metric_name_raw = str(row.get("metric_name", "")).strip()
        metric_name = _normalize_metric_name(metric_name_raw)

        predicted_value = _safe_float(row.get("predicted_value"), np.nan)
        replayed_value = _safe_float(row.get("replayed_value"), np.nan)
        limit_low = _safe_float(row.get("limit_low"), np.nan)
        limit_high = _safe_float(row.get("limit_high"), np.nan)

        low = limit_low
        high = limit_high

        if metric_name_raw.startswith("Delta_P_IBR_"):
            if delta_p_mode == "predicted headroom":
                up = _safe_float(row.get("scheduled_headroom_up"), np.nan)
                down = _safe_float(row.get("scheduled_headroom_down"), np.nan)
                low = -down if np.isfinite(down) else np.nan
                high = up if np.isfinite(up) else np.nan
            elif delta_p_mode == "physical headroom":
                low = _safe_float(row.get("physical_min"), np.nan)
                high = _safe_float(row.get("physical_max"), np.nan)

        within = int(np.isfinite(replayed_value) and np.isfinite(low) and np.isfinite(high) and (replayed_value >= low - 1e-8) and (replayed_value <= high + 1e-8))
        if delta_p_mode == "existing_repo_mode" and metric_name_raw.startswith("Delta_P_IBR_"):
            existing_flag = _safe_float(row.get("replayed_within_limits"), np.nan)
            if np.isfinite(existing_flag):
                within = int(existing_flag >= 0.5)

        limit_value = np.nan
        if np.isfinite(low) and np.isfinite(high):
            if metric_name_raw.startswith("Delta_P_IBR_"):
                if np.isfinite(replayed_value):
                    limit_value = high if replayed_value >= 0 else abs(low)
                else:
                    limit_value = max(abs(low), abs(high))
            else:
                limit_value = max(abs(low), abs(high))

        out_rows.append(
            {
                "formulation_code": formulation_code,
                "metric_name": metric_name,
                "predicted_value": predicted_value,
                "replayed_value": replayed_value,
                "limit_value": limit_value,
                "is_within_limit": within,
                "case_label": case_label,
                "bounds_case": bounds_case,
                "contingency_label": contingency_label,
            }
        )

    return pd.DataFrame(out_rows, columns=columns)


def _apply_time_window_and_sampling(df: pd.DataFrame, logging_cfg: dict[str, Any]) -> pd.DataFrame:
    out = df.copy()
    window = logging_cfg.get("trajectory_time_window_s")
    if isinstance(window, (list, tuple)) and len(window) == 2:
        t0 = float(window[0])
        t1 = float(window[1])
        out = out.loc[(pd.to_numeric(out.get("time_s"), errors="coerce") >= t0) & (pd.to_numeric(out.get("time_s"), errors="coerce") <= t1)].copy()

    step = _safe_float(logging_cfg.get("trajectory_sampling_step"), np.nan)
    if np.isfinite(step) and step > 0:
        out = out.sort_values("time_s").copy()
        keep_mask = np.zeros(len(out), dtype=bool)
        last_t = None
        t_values = pd.to_numeric(out["time_s"], errors="coerce").to_numpy(dtype=float)
        for i, t in enumerate(t_values):
            if not np.isfinite(t):
                continue
            if last_t is None or (t - last_t) >= step - 1e-9:
                keep_mask[i] = True
                last_t = float(t)
        out = out.loc[keep_mask].copy()
    return out


def _collect_trajectory_tables(
    *,
    trace_summary_csv: Path,
    map_id_to_code: dict[str, str],
    per_formulation_ibr_map: dict[str, dict[int, dict[str, Any]]],
    cfg: dict[str, Any],
    case_label: str,
    bounds_case: str,
    sbase: float,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    coi_columns = [
        "time_s",
        "formulation_code",
        "bounds_case",
        "delta_f_coi_hz",
        "rocof_coi_hz_per_s",
        "f_coi_hz",
        "case_label",
    ]
    ibr_columns = [
        "time_s",
        "formulation_code",
        "bounds_case",
        "unit_name",
        "delta_p_ibr_mw",
        "p_ibr_mw",
        "case_label",
    ]
    ibr_pu_columns = [
        "time_s",
        "formulation_code",
        "bounds_case",
        "unit_name",
        "delta_p_ibr_pu",
        "p_ibr_pu",
        "case_label",
    ]

    if not trace_summary_csv.exists():
        return pd.DataFrame(columns=coi_columns), pd.DataFrame(columns=ibr_columns), pd.DataFrame(columns=ibr_pu_columns)

    summary_df = pd.read_csv(trace_summary_csv)
    if summary_df.empty or "trace_csv" not in summary_df.columns:
        return pd.DataFrame(columns=coi_columns), pd.DataFrame(columns=ibr_columns), pd.DataFrame(columns=ibr_pu_columns)

    logging_cfg = dict(cfg.get("logging", {}) or {})

    coi_rows: list[dict[str, Any]] = []
    ibr_rows: list[dict[str, Any]] = []
    ibr_rows_pu: list[dict[str, Any]] = []

    for _, row in summary_df.iterrows():
        trace_csv = Path(str(row.get("trace_csv", "")).strip())
        if not trace_csv.exists():
            continue

        trace_df = pd.read_csv(trace_csv)
        if trace_df.empty:
            continue

        formulation_id = str(row.get("formulation_id", "")).strip()
        if not formulation_id and "formulation_id" in trace_df.columns:
            formulation_id = str(trace_df["formulation_id"].dropna().iloc[0]) if not trace_df["formulation_id"].dropna().empty else ""
        formulation_code = map_id_to_code.get(formulation_id, FORMULATION_CODE_BY_ID.get(formulation_id, formulation_id or str(row.get("label", ""))))

        if not bool(logging_cfg.get("save_time_series", True)):
            continue

        per_trace = trace_df.copy()
        per_trace = _apply_time_window_and_sampling(per_trace, logging_cfg)

        for _, tr in per_trace.iterrows():
            coi_rows.append(
                {
                    "time_s": _safe_float(tr.get("time_s"), np.nan),
                    "formulation_code": formulation_code,
                    "bounds_case": bounds_case,
                    "delta_f_coi_hz": _safe_float(tr.get("delta_f_coi_hz"), np.nan),
                    "rocof_coi_hz_per_s": _safe_float(tr.get("rocof_coi_hz_per_s"), np.nan),
                    "f_coi_hz": _safe_float(tr.get("f_coi_hz"), np.nan),
                    "case_label": case_label,
                }
            )

        ibr_map = per_formulation_ibr_map.get(formulation_code, {})
        ibr_cols = [col for col in per_trace.columns if re.match(r"^Delta_P_IBR_\d+$", str(col))]
        for col in ibr_cols:
            idx = int(str(col).split("_")[-1])
            unit_meta = dict(ibr_map.get(idx, {}) or {})
            unit_name = str(unit_meta.get("unit_name", f"ibr_{idx}"))
            scheduled_pu = _safe_float(unit_meta.get("scheduled_dispatch_pu"), np.nan)

            series = pd.to_numeric(per_trace[col], errors="coerce")
            time_series = pd.to_numeric(per_trace["time_s"], errors="coerce")
            for t, delta_p_pu in zip(time_series.to_numpy(dtype=float), series.to_numpy(dtype=float)):
                p_ibr_pu = scheduled_pu + delta_p_pu if np.isfinite(scheduled_pu) and np.isfinite(delta_p_pu) else np.nan
                ibr_rows.append(
                    {
                        "time_s": t,
                        "formulation_code": formulation_code,
                        "bounds_case": bounds_case,
                        "unit_name": unit_name,
                        "delta_p_ibr_mw": delta_p_pu * sbase if np.isfinite(delta_p_pu) else np.nan,
                        "p_ibr_mw": p_ibr_pu * sbase if np.isfinite(p_ibr_pu) else np.nan,
                        "case_label": case_label,
                    }
                )
                ibr_rows_pu.append(
                    {
                        "time_s": t,
                        "formulation_code": formulation_code,
                        "bounds_case": bounds_case,
                        "unit_name": unit_name,
                        "delta_p_ibr_pu": delta_p_pu,
                        "p_ibr_pu": p_ibr_pu,
                        "case_label": case_label,
                    }
                )

    return (
        pd.DataFrame(coi_rows, columns=coi_columns),
        pd.DataFrame(ibr_rows, columns=ibr_columns),
        pd.DataFrame(ibr_rows_pu, columns=ibr_pu_columns),
    )


def _write_run_readme(
    *,
    out_path: Path,
    case_label: str,
    bounds_case: str,
    csv_notes: list[str],
    missing_notes: list[str],
    generated_files: list[str],
) -> None:
    lines: list[str] = []
    lines.append(f"# Presentation VIS Run Summary: {case_label} ({bounds_case})")
    lines.append("")
    lines.append("## What was run")
    lines.append("- Optimization runs were executed via `scheduling/run_experiment_suite.py`.")
    lines.append("- Replay validation was executed via `scheduling/replay_validation.py` when enabled.")
    lines.append("- Replay trajectories were exported via `results/thesis_optimization_results/scripts/export_replay_trace_panel.py` when enabled.")
    lines.append("")
    lines.append("## Output CSVs")
    lines.append("- `formulations_summary.csv`: formulation flags and solve status/timing.")
    lines.append("- `costs_by_formulation.csv`: optimization cost decomposition from summary JSON objective components.")
    lines.append("- `setpoints_by_formulation.csv`: dispatch/headroom/reserve and VIS M/D per unit.")
    lines.append("- `replay_metrics_by_formulation.csv`: predicted vs replayed metrics and limit membership.")
    lines.append("- `coi_trajectories_by_formulation.csv`: COI frequency/RoCoF trajectories.")
    lines.append("- `ibr_power_trajectories_by_formulation.csv`: IBR power trajectories in MW.")
    lines.append("- `ibr_power_trajectories_by_formulation_pu.csv`: IBR power trajectories in p.u. (if enabled).")
    lines.append("")
    if csv_notes:
        lines.append("## Notes")
        for note in csv_notes:
            lines.append(f"- {note}")
        lines.append("")

    if missing_notes:
        lines.append("## Missing/Approximated")
        for note in missing_notes:
            lines.append(f"- {note}")
        lines.append("")

    lines.append("## Generated Files")
    for path in generated_files:
        lines.append(f"- `{path}`")
    lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")


def _export_case_outputs(
    *,
    cfg: dict[str, Any],
    out_dir: Path,
    map_id_to_code: dict[str, str],
    suite_summary_json: Path,
    replay_detail_csv: Path,
    trace_summary_csv: Path,
    metadata_extra: dict[str, Any],
) -> dict[str, Any]:
    case_label = str(cfg["run"]["case_label"])
    bounds_case = str(out_dir.name)
    sbase = float(cfg.get("system", {}).get("sbase", 100.0))

    rows = _load_suite_rows(suite_summary_json)
    formulations_df, costs_df, setpoints_df, per_formulation_ibr_map, errors = _collect_optimization_tables(
        rows=rows,
        map_id_to_code=map_id_to_code,
        case_label=case_label,
        bounds_case=bounds_case,
        sbase=sbase,
    )

    replay_df = _collect_replay_metrics(
        replay_detail_csv=replay_detail_csv,
        map_id_to_code=map_id_to_code,
        cfg=cfg,
        case_label=case_label,
        bounds_case=bounds_case,
    )

    coi_df, ibr_df, ibr_pu_df = _collect_trajectory_tables(
        trace_summary_csv=trace_summary_csv,
        map_id_to_code=map_id_to_code,
        per_formulation_ibr_map=per_formulation_ibr_map,
        cfg=cfg,
        case_label=case_label,
        bounds_case=bounds_case,
        sbase=sbase,
    )

    formulations_cols = [
        "formulation_code",
        "formulation_name",
        "includes_line_constraints",
        "includes_n1_redispatch",
        "includes_surrogate",
        "includes_vis",
        "is_feasible",
        "solve_time_ms",
        "case_label",
        "bounds_case",
    ]
    costs_cols = [
        "formulation_code",
        "objective_cost_kusd",
        "dispatch_only_cost_kusd",
        "reserve_cost_kusd",
        "other_cost_component",
        "solve_time_ms",
        "case_label",
        "bounds_case",
    ]
    setpoints_cols = [
        "formulation_code",
        "unit_name",
        "unit_type",
        "scheduled_dispatch_mw",
        "upward_headroom_mw",
        "downward_headroom_mw",
        "virtual_inertia",
        "virtual_damping",
        "reserve_up_mw",
        "reserve_down_mw",
        "case_label",
        "bounds_case",
    ]
    replay_cols = [
        "formulation_code",
        "metric_name",
        "predicted_value",
        "replayed_value",
        "limit_value",
        "is_within_limit",
        "case_label",
        "bounds_case",
        "contingency_label",
    ]
    coi_cols = [
        "time_s",
        "formulation_code",
        "bounds_case",
        "delta_f_coi_hz",
        "rocof_coi_hz_per_s",
        "f_coi_hz",
        "case_label",
    ]
    ibr_cols = [
        "time_s",
        "formulation_code",
        "bounds_case",
        "unit_name",
        "delta_p_ibr_mw",
        "p_ibr_mw",
        "case_label",
    ]

    _write_df(out_dir / "formulations_summary.csv", formulations_df, formulations_cols)
    _write_df(out_dir / "costs_by_formulation.csv", costs_df, costs_cols)
    _write_df(out_dir / "setpoints_by_formulation.csv", setpoints_df, setpoints_cols)
    _write_df(out_dir / "replay_metrics_by_formulation.csv", replay_df, replay_cols)
    _write_df(out_dir / "coi_trajectories_by_formulation.csv", coi_df, coi_cols)
    _write_df(out_dir / "ibr_power_trajectories_by_formulation.csv", ibr_df, ibr_cols)

    generated = [
        str(out_dir / "formulations_summary.csv"),
        str(out_dir / "costs_by_formulation.csv"),
        str(out_dir / "setpoints_by_formulation.csv"),
        str(out_dir / "replay_metrics_by_formulation.csv"),
        str(out_dir / "coi_trajectories_by_formulation.csv"),
        str(out_dir / "ibr_power_trajectories_by_formulation.csv"),
    ]

    if bool(cfg.get("logging", {}).get("save_per_unit_and_mw_versions", True)):
        _write_df(
            out_dir / "ibr_power_trajectories_by_formulation_pu.csv",
            ibr_pu_df,
            [
                "time_s",
                "formulation_code",
                "bounds_case",
                "unit_name",
                "delta_p_ibr_pu",
                "p_ibr_pu",
                "case_label",
            ],
        )
        generated.append(str(out_dir / "ibr_power_trajectories_by_formulation_pu.csv"))

    missing_notes = list(errors)
    if not replay_detail_csv.exists():
        missing_notes.append("Replay detail CSV is missing (replay disabled or failed), so replay_metrics_by_formulation.csv may be empty.")
    if not trace_summary_csv.exists():
        missing_notes.append("Trace summary CSV is missing (time-series export disabled or failed), so trajectory CSVs may be empty.")

    csv_notes = [
        "Cost decomposition follows existing optimization summary fields (`objective_dispatch_only`, `objective_reserve_only`, tie-breaker/regularization terms).",
        "Post-contingency reserve (`objective_reserve_postcont_only`) is excluded from `reserve_cost_kusd` and `objective_cost_kusd` when available.",
        "`objective_cost_kusd` is computed from existing optimization components without extra scaling.",
        "IBR trajectory MW values are derived from p.u. traces using `system.sbase`.",
        "`delta_p_mode` is applied in export-time limit checks for replay metrics; core optimization constraints remain the repository default behavior.",
    ]

    metadata = {
        "timestamp": _now_iso(),
        "git_commit_hash": _git_commit_hash(),
        "path_to_yaml": str(metadata_extra.get("config_path", "")),
        "exact_command_run": metadata_extra.get("exact_command_run", ""),
        "executed_subcommands": metadata_extra.get("executed_subcommands", []),
        "case_label": case_label,
        "bounds_case": bounds_case,
        "model_bundle_used": str(cfg.get("surrogate", {}).get("model_bundle_path", "")),
        "scaler_bundle_used": str(cfg.get("surrogate", {}).get("scaler_bundle_path", "")),
        "disturbance_settings": dict(cfg.get("system", {}) or {}),
        "contingency_settings": {
            "contingency_label": cfg.get("system", {}).get("contingency_label"),
            "contingency_mode": cfg.get("system", {}).get("contingency_mode"),
            "topk_contingencies": cfg.get("network_security", {}).get("topk_contingencies"),
            "screened_contingency_set_name": cfg.get("network_security", {}).get("screened_contingency_set_name"),
        },
        "security_limits": dict(cfg.get("security_limits", {}) or {}),
        "formulations_run": sorted(set(formulations_df.get("formulation_code", pd.Series(dtype=str)).astype(str).tolist())),
        "replay_succeeded": bool(metadata_extra.get("replay_succeeded", False)),
        "notes": missing_notes,
    }
    _write_json(out_dir / "metadata.json", metadata)

    generated += [str(out_dir / "metadata.json")]

    _write_run_readme(
        out_path=out_dir / "README_run.md",
        case_label=case_label,
        bounds_case=bounds_case,
        csv_notes=csv_notes,
        missing_notes=missing_notes,
        generated_files=generated,
    )
    generated.append(str(out_dir / "README_run.md"))

    return {
        "generated_files": generated,
        "formulations_summary": str(out_dir / "formulations_summary.csv"),
        "costs_by_formulation": str(out_dir / "costs_by_formulation.csv"),
        "setpoints_by_formulation": str(out_dir / "setpoints_by_formulation.csv"),
    }


def _build_schedule_diff(case_root: Path, case_label: str, default_suffix: str, tight_suffix: str) -> Path:
    default_dir = case_root / default_suffix
    tight_dir = case_root / tight_suffix
    out_path = case_root / "schedule_diff_default_vs_tight.csv"

    set_cols = [
        "formulation_code",
        "unit_name",
        "scheduled_dispatch_mw",
        "upward_headroom_mw",
        "virtual_inertia",
        "virtual_damping",
    ]
    cost_cols = [
        "formulation_code",
        "objective_cost_kusd",
        "dispatch_only_cost_kusd",
        "reserve_cost_kusd",
    ]

    if not (default_dir / "setpoints_by_formulation.csv").exists() or not (tight_dir / "setpoints_by_formulation.csv").exists():
        pd.DataFrame(
            columns=[
                "formulation_code",
                "unit_name",
                "dispatch_diff_mw",
                "upward_headroom_diff_mw",
                "virtual_inertia_diff",
                "virtual_damping_diff",
                "objective_cost_diff_kusd",
                "dispatch_cost_diff_kusd",
                "reserve_cost_diff_kusd",
                "case_label",
            ]
        ).to_csv(out_path, index=False)
        return out_path

    default_set = pd.read_csv(default_dir / "setpoints_by_formulation.csv")
    tight_set = pd.read_csv(tight_dir / "setpoints_by_formulation.csv")
    default_set = default_set[[col for col in set_cols if col in default_set.columns]].copy()
    tight_set = tight_set[[col for col in set_cols if col in tight_set.columns]].copy()

    merged = default_set.merge(
        tight_set,
        on=["formulation_code", "unit_name"],
        how="inner",
        suffixes=("_default", "_tight"),
    )

    if (default_dir / "costs_by_formulation.csv").exists() and (tight_dir / "costs_by_formulation.csv").exists():
        default_cost = pd.read_csv(default_dir / "costs_by_formulation.csv")
        tight_cost = pd.read_csv(tight_dir / "costs_by_formulation.csv")
        default_cost = default_cost[[col for col in cost_cols if col in default_cost.columns]].copy()
        tight_cost = tight_cost[[col for col in cost_cols if col in tight_cost.columns]].copy()
        cost_merged = default_cost.merge(tight_cost, on=["formulation_code"], how="inner", suffixes=("_default", "_tight"))
        for col in ["objective_cost_kusd", "dispatch_only_cost_kusd", "reserve_cost_kusd"]:
            cost_merged[f"{col}_diff"] = pd.to_numeric(cost_merged.get(f"{col}_tight"), errors="coerce") - pd.to_numeric(
                cost_merged.get(f"{col}_default"), errors="coerce"
            )
        merged = merged.merge(
            cost_merged[["formulation_code", "objective_cost_kusd_diff", "dispatch_only_cost_kusd_diff", "reserve_cost_kusd_diff"]],
            on="formulation_code",
            how="left",
        )
    else:
        merged["objective_cost_kusd_diff"] = np.nan
        merged["dispatch_only_cost_kusd_diff"] = np.nan
        merged["reserve_cost_kusd_diff"] = np.nan

    out_df = pd.DataFrame(
        {
            "formulation_code": merged.get("formulation_code"),
            "unit_name": merged.get("unit_name"),
            "dispatch_diff_mw": pd.to_numeric(merged.get("scheduled_dispatch_mw_tight"), errors="coerce") - pd.to_numeric(
                merged.get("scheduled_dispatch_mw_default"), errors="coerce"
            ),
            "upward_headroom_diff_mw": pd.to_numeric(merged.get("upward_headroom_mw_tight"), errors="coerce") - pd.to_numeric(
                merged.get("upward_headroom_mw_default"), errors="coerce"
            ),
            "virtual_inertia_diff": pd.to_numeric(merged.get("virtual_inertia_tight"), errors="coerce") - pd.to_numeric(
                merged.get("virtual_inertia_default"), errors="coerce"
            ),
            "virtual_damping_diff": pd.to_numeric(merged.get("virtual_damping_tight"), errors="coerce") - pd.to_numeric(
                merged.get("virtual_damping_default"), errors="coerce"
            ),
            "objective_cost_diff_kusd": pd.to_numeric(merged.get("objective_cost_kusd_diff"), errors="coerce"),
            "dispatch_cost_diff_kusd": pd.to_numeric(merged.get("dispatch_only_cost_kusd_diff"), errors="coerce"),
            "reserve_cost_diff_kusd": pd.to_numeric(merged.get("reserve_cost_kusd_diff"), errors="coerce"),
            "case_label": case_label,
        }
    )

    out_df.to_csv(out_path, index=False)
    return out_path


def run_pipeline(config_path: Path) -> dict[str, Any]:
    raw_cfg = _load_yaml(config_path)
    cfg = _apply_defaults(raw_cfg)
    _validate_config(cfg)

    run_cfg = dict(cfg.get("run", {}) or {})
    case_label = str(cfg["run"]["case_label"]).strip()
    output_root = _resolve(str(cfg["run"]["output_root"])).resolve()
    generated_configs_root_raw = str(
        run_cfg.get("generated_configs_root", os.environ.get("PRESENTATION_VIS_GENERATED_CONFIGS_ROOT", ""))
    ).strip()
    generated_configs_root = _resolve(generated_configs_root_raw).resolve() if generated_configs_root_raw else None
    case_root = output_root / case_label
    case_root.mkdir(parents=True, exist_ok=True)

    comparison_cfg = dict(cfg.get("comparison", {}) or {})
    default_suffix = str(comparison_cfg.get("default_case_suffix", "default"))
    tight_suffix = str(comparison_cfg.get("tight_case_suffix", "tight"))

    run_cases: list[tuple[str, bool]] = []
    run_default = bool(comparison_cfg.get("run_default_bounds_case", False))
    run_tight = bool(comparison_cfg.get("run_tight_bounds_case", False))

    if run_default:
        apply_tight_default = (not run_tight)
        run_cases.append((default_suffix, apply_tight_default))
    if run_tight:
        run_cases.append((tight_suffix, True))

    python_bin = os.environ.get("PYTHON_BIN", sys.executable)
    env = os.environ.copy()

    summary: dict[str, Any] = {
        "case_label": case_label,
        "config_path": str(config_path),
        "case_root": str(case_root),
        "bounds_cases": {},
    }

    executed_subcommands: list[list[str]] = []
    replay_succeeded_overall = True

    for bounds_case, apply_tight_limits in run_cases:
        out_dir = case_root / bounds_case
        out_dir.mkdir(parents=True, exist_ok=True)
        raw_root = out_dir / "raw"
        if generated_configs_root is None:
            generated_cfg_root = raw_root / "generated_configs"
        else:
            generated_cfg_root = generated_configs_root / case_label / bounds_case
        generated_cfg_root.mkdir(parents=True, exist_ok=True)

        suite_cfg, map_id_to_code, notes, replay_contingency = _build_suite_config(
            cfg,
            bounds_case=bounds_case,
            out_root=out_dir,
            apply_tight_limits=apply_tight_limits,
        )

        suite_cfg_path = generated_cfg_root / "suite.yaml"
        _write_yaml(suite_cfg_path, suite_cfg)

        log_tail_lines = int(_safe_float(cfg.get("logging", {}).get("log_tail_lines"), 120.0))
        cmd_suite = [
            python_bin,
            "scheduling/run_experiment_suite.py",
            "--suite",
            str(suite_cfg_path),
            "--log-tail-lines",
            str(log_tail_lines),
        ]
        _run_command(cmd_suite, cwd=ROOT, env=env)
        executed_subcommands.append(cmd_suite)

        suite_summary_json = Path(str(suite_cfg["output"]["summary_json"]))
        suite_rows = _load_suite_rows(suite_summary_json)
        formulation_ids = sorted({str(row.get("formulation_id", "")).strip() for row in suite_rows if str(row.get("formulation_id", "")).strip()})

        replay_detail_csv = raw_root / "replay_validation_detail.csv"
        trace_summary_csv = raw_root / "traces" / "trace_summary.csv"

        if bool(cfg.get("run", {}).get("run_replay", True)):
            replay_cfg = {
                "tds": {
                    "t_end": 21.0,
                    "t_step": 0.1,
                    "no_tqdm": True,
                    "criteria": 0,
                    "tol": 1.0e-3,
                    "fixt": 0,
                    "method": "backeuler",
                    "honest": 0,
                    "max_iter": 35,
                    "shrinkt": 1,
                },
                "contingency": replay_contingency,
                "output_metric_map": {
                    "rocof_COI": "coi.rocof_COI",
                    "dev_COI": "coi.dev_COI",
                    "Delta_P_IBR_1": "ibr_response.Delta_P_IBR_1",
                    "Delta_P_IBR_2": "ibr_response.Delta_P_IBR_2",
                    "Delta_P_IBR_3": "ibr_response.Delta_P_IBR_3",
                    "Delta_P_IBR_4": "ibr_response.Delta_P_IBR_4",
                },
                "runs": [
                    {
                        "suite_summary_json": str(suite_summary_json),
                        "formulation_ids": formulation_ids,
                    }
                ],
                "output": {
                    "summary_csv": str(raw_root / "replay_validation_summary.csv"),
                    "detail_csv": str(replay_detail_csv),
                    "summary_json": str(raw_root / "replay_validation_summary.json"),
                },
            }
            replay_cfg_path = generated_cfg_root / "replay_validation.yaml"
            _write_yaml(replay_cfg_path, replay_cfg)

            cmd_replay = [
                python_bin,
                "scheduling/replay_validation.py",
                "--config",
                str(replay_cfg_path),
            ]
            try:
                _run_command(cmd_replay, cwd=ROOT, env=env)
                executed_subcommands.append(cmd_replay)
            except Exception:
                replay_succeeded_overall = False

            trace_runs = []
            for row in suite_rows:
                status = str(row.get("status", "")).strip().lower()
                if not status.startswith("optimal"):
                    continue
                formulation_id = str(row.get("formulation_id", "")).strip()
                code = map_id_to_code.get(formulation_id, FORMULATION_CODE_BY_ID.get(formulation_id, formulation_id))
                label = FORMULATION_LABEL_BY_ID.get(formulation_id, code)
                summary_json = str(row.get("summary_json", "")).strip()
                if summary_json:
                    trace_runs.append(
                        {
                            "label": label,
                            "summary_json": summary_json,
                        }
                    )

            if bool(cfg.get("formulations", {}).get("run_no_vi_reference", False)) and trace_runs:
                trace_runs.append(
                    {
                        "label": "No VI reference",
                        "summary_json": trace_runs[0]["summary_json"],
                        "linestyle": ":",
                        "md_override": {
                            "m": [0.0, 0.0, 0.0, 0.0],
                            "d": [0.0, 0.0, 0.0, 0.0],
                        },
                    }
                )

            if trace_runs:
                trace_cfg = {
                    "tds": {
                        "t_end": 20.0,
                        "t_step": 0.01,
                        "method": "backeuler",
                        "no_tqdm": True,
                        "criteria": 0,
                    },
                    "contingency": replay_contingency,
                    "limits": {
                        "frequency_hz": 50.0,
                        "delta_f_hz": float(cfg.get("security_limits", {}).get("delta_f_lim_hz", 0.8)),
                        "rocof_hz_per_s": float(cfg.get("security_limits", {}).get("rocof_lim_hz_per_s", 1.0)),
                    },
                    "output": {
                        "directory": str(raw_root / "traces"),
                        "plot_mode": "combined",
                        "ibr_scale": float(cfg.get("system", {}).get("sbase", 100.0)),
                        "ibr_unit": "MW",
                    },
                    "runs": trace_runs,
                }
                trace_cfg_path = generated_cfg_root / "replay_trace.yaml"
                _write_yaml(trace_cfg_path, trace_cfg)

                cmd_trace = [
                    python_bin,
                    "results/thesis_optimization_results/scripts/export_replay_trace_panel.py",
                    "--config",
                    str(trace_cfg_path),
                ]
                try:
                    _run_command(cmd_trace, cwd=ROOT, env=env)
                    executed_subcommands.append(cmd_trace)
                except Exception:
                    replay_succeeded_overall = False

        exported = _export_case_outputs(
            cfg=cfg,
            out_dir=out_dir,
            map_id_to_code=map_id_to_code,
            suite_summary_json=suite_summary_json,
            replay_detail_csv=replay_detail_csv,
            trace_summary_csv=trace_summary_csv,
            metadata_extra={
                "config_path": str(config_path),
                "exact_command_run": " ".join([sys.executable, *sys.argv]),
                "executed_subcommands": executed_subcommands,
                "replay_succeeded": replay_succeeded_overall,
            },
        )

        summary["bounds_cases"][bounds_case] = {
            "output_dir": str(out_dir),
            "generated_config_dir": str(generated_cfg_root),
            "suite_summary_json": str(suite_summary_json),
            "notes": notes,
            "exported": exported,
        }

    if run_default and run_tight:
        diff_path = _build_schedule_diff(case_root, case_label, default_suffix=default_suffix, tight_suffix=tight_suffix)
        summary["schedule_diff_default_vs_tight"] = str(diff_path)

    _write_json(case_root / "run_summary.json", summary)
    return summary


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Run a presentation VIS case and export slide-ready CSV outputs.")
    parser.add_argument("--config", required=True, help="Path to presentation VIS YAML config.")
    args = parser.parse_args(argv)

    config_path = _resolve(args.config)
    if not config_path.exists():
        raise PipelineError(f"Config file not found: {config_path}")

    summary = run_pipeline(config_path)
    print("[presentation_vis] Completed run.")
    print(f"[presentation_vis] Case root: {summary['case_root']}")
    if "schedule_diff_default_vs_tight" in summary:
        print(f"[presentation_vis] Diff CSV: {summary['schedule_diff_default_vs_tight']}")


if __name__ == "__main__":
    main()
