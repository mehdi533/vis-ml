from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import pandas as pd
from matplotlib import pyplot as plt

ROOT = Path(__file__).resolve().parents[3]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    from results.thesis_optimization_results.src.analysis_utils import (
        write_latex_table,
        write_markdown_table,
    )
    from results.thesis_optimization_results.src.plot_utils import (
        formulation_color,
        save_figure,
        set_thesis_style,
    )
    from results.thesis_optimization_results.src.config_analysis import (
        FIGURES_DIR,
        TABLES_DIR,
        ensure_output_dirs,
    )
except ImportError:
    from analysis_utils import write_latex_table, write_markdown_table  # type: ignore
    from plot_utils import formulation_color, save_figure, set_thesis_style  # type: ignore
    from config_analysis import FIGURES_DIR, TABLES_DIR, ensure_output_dirs  # type: ignore


def _write_table_bundle(stem: str, df: pd.DataFrame) -> dict[str, str]:
    csv_path = TABLES_DIR / f"{stem}.csv"
    md_path = TABLES_DIR / f"{stem}.md"
    tex_path = TABLES_DIR / f"{stem}.tex"
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(csv_path, index=False)
    write_markdown_table(md_path, df)
    write_latex_table(tex_path, df)
    return {"csv": str(csv_path), "markdown": str(md_path), "latex": str(tex_path)}


def _load_suite_summary(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _load_area_map(case_path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(case_path, data_only=True, read_only=True)

    ws_area = wb["Area"]
    area_rows = list(ws_area.iter_rows(values_only=True))
    area_head = area_rows[0]
    area_name_by_idx: dict[int, str] = {}
    for row in area_rows[1:]:
        if not row or row[0] is None:
            continue
        payload = dict(zip(area_head, row))
        area_name_by_idx[int(payload["idx"])] = str(payload["name"])

    ws_bus = wb["Bus"]
    bus_rows = list(ws_bus.iter_rows(values_only=True))
    bus_head = bus_rows[0]
    bus_meta: dict[int, dict[str, Any]] = {}
    for row in bus_rows[1:]:
        if not row or row[0] is None:
            continue
        payload = dict(zip(bus_head, row))
        bus_idx = int(payload["idx"])
        area_idx = int(payload["area"])
        bus_meta[bus_idx] = {
            "bus_idx": bus_idx,
            "bus_name": str(payload["name"]),
            "area_idx": area_idx,
            "area_name": area_name_by_idx.get(area_idx, str(area_idx)),
            "zone_idx": int(payload["zone"]) if payload.get("zone") is not None else np.nan,
        }

    ws_regcv1 = wb["REGCV1"]
    reg_rows = list(ws_regcv1.iter_rows(values_only=True))
    reg_head = reg_rows[0]
    out_rows = []
    local_index = 0
    for row in reg_rows[1:]:
        if not row or row[0] is None:
            continue
        payload = dict(zip(reg_head, row))
        if int(payload.get("u", 1)) != 1:
            continue
        bus_idx = int(payload["bus"])
        meta = dict(bus_meta.get(bus_idx, {}))
        out_rows.append(
            {
                "index": int(local_index),
                "regcv1_uid": int(payload["uid"]),
                "regcv1_idx": str(payload["idx"]),
                "regcv1_name": str(payload["name"]),
                "bus_idx": bus_idx,
                "bus_name": meta.get("bus_name", f"BUS{bus_idx}"),
                "area_idx": meta.get("area_idx", np.nan),
                "area_name": meta.get("area_name", "UNKNOWN"),
                "zone_idx": meta.get("zone_idx", np.nan),
            }
        )
        local_index += 1
    return pd.DataFrame(out_rows)


def _load_ibr_rows(summary_json: Path) -> tuple[pd.DataFrame, str | None]:
    payload = _load_suite_summary(summary_json)
    rows = []
    system_case: str | None = None

    for row in list(payload.get("rows") or []):
        summary_path_raw = str(row.get("summary_json", "")).strip()
        if not summary_path_raw:
            continue
        summary_path = Path(summary_path_raw)
        if not summary_path.exists():
            continue
        with summary_path.open("r", encoding="utf-8") as f:
            summary_payload = json.load(f)

        artifacts = dict(summary_payload.get("artifacts", {}) or {})
        dispatch_path_raw = str(artifacts.get("dispatch_impact_csv", "")).strip()
        if not dispatch_path_raw:
            continue
        dispatch_path = Path(dispatch_path_raw)
        if not dispatch_path.exists():
            continue

        if system_case is None:
            system_case = str(summary_payload.get("system_case", "")).strip() or None

        dispatch_df = pd.read_csv(dispatch_path)
        dispatch_df.columns = [str(col).strip() for col in dispatch_df.columns]
        if "row_type" not in dispatch_df.columns:
            continue
        dispatch_df = dispatch_df.loc[dispatch_df["row_type"].astype(str) == "ibr_summary"].copy()
        if dispatch_df.empty:
            continue
        dispatch_df["run_id"] = str(summary_payload.get("run_id", row.get("run_id", "")))
        dispatch_df["formulation_id"] = str(summary_payload.get("formulation_id", row.get("formulation_id", "")))
        dispatch_df["formulation_name"] = str(summary_payload.get("formulation_name", row.get("formulation_name", "")))
        dispatch_df["paper_method"] = str(row.get("paper_method", ""))
        dispatch_df["comparison_family"] = str(row.get("comparison_family", ""))
        dispatch_df["scenario_id"] = str(summary_payload.get("scenario_id", row.get("scenario_id", "")))
        dispatch_df["scenario_name"] = str(summary_payload.get("scenario_name", row.get("scenario_name", "")))
        dispatch_df["status"] = str(summary_payload.get("status", row.get("status", "")))
        dispatch_df["objective_dispatch_only"] = summary_payload.get("objective_dispatch_only", row.get("objective_dispatch_only"))
        dispatch_df["solve_time_sec"] = dict(summary_payload.get("solver_stats", {}) or {}).get("solve_time_sec", row.get("solve_time_sec"))
        rows.append(dispatch_df)

    out = pd.concat(rows, ignore_index=True, sort=False) if rows else pd.DataFrame()
    return out, system_case


def _area_summary(ibr_df: pd.DataFrame, *, baseline_id: str) -> pd.DataFrame:
    if ibr_df.empty:
        return ibr_df.copy()

    numeric_cols = [
        "M_opt",
        "D_opt",
        "delta_p_dispatch",
        "delta_p_predicted",
        "headroom_up",
        "headroom_down",
    ]
    for col in numeric_cols:
        ibr_df[col] = pd.to_numeric(ibr_df[col], errors="coerce")

    grouped = (
        ibr_df.groupby(
            ["scenario_id", "scenario_name", "formulation_id", "formulation_name", "paper_method", "comparison_family", "area_name"],
            dropna=False,
        )
        .agg(
            n_ibr=("index", "count"),
            M_sum=("M_opt", "sum"),
            M_mean=("M_opt", "mean"),
            D_sum=("D_opt", "sum"),
            D_mean=("D_opt", "mean"),
            delta_p_dispatch_sum=("delta_p_dispatch", "sum"),
            delta_p_predicted_sum=("delta_p_predicted", "sum"),
            headroom_up_sum=("headroom_up", "sum"),
            headroom_down_sum=("headroom_down", "sum"),
        )
        .reset_index()
    )

    totals = (
        grouped.groupby(["scenario_id", "formulation_id"], dropna=False)
        .agg(total_M_sum=("M_sum", "sum"), total_D_sum=("D_sum", "sum"))
        .reset_index()
    )
    grouped = grouped.merge(totals, on=["scenario_id", "formulation_id"], how="left")
    grouped["M_area_share"] = grouped["M_sum"] / grouped["total_M_sum"].replace({0.0: np.nan})
    grouped["D_area_share"] = grouped["D_sum"] / grouped["total_D_sum"].replace({0.0: np.nan})

    baseline = (
        grouped.loc[grouped["formulation_id"].astype(str) == str(baseline_id)]
        .rename(
            columns={
                "M_sum": "baseline_M_sum",
                "M_mean": "baseline_M_mean",
                "D_sum": "baseline_D_sum",
                "D_mean": "baseline_D_mean",
                "delta_p_dispatch_sum": "baseline_delta_p_dispatch_sum",
                "headroom_up_sum": "baseline_headroom_up_sum",
            }
        )
        .copy()
    )
    keep = [
        "scenario_id",
        "area_name",
        "baseline_M_sum",
        "baseline_M_mean",
        "baseline_D_sum",
        "baseline_D_mean",
        "baseline_delta_p_dispatch_sum",
        "baseline_headroom_up_sum",
    ]
    grouped = grouped.merge(baseline[keep], on=["scenario_id", "area_name"], how="left")
    grouped["M_sum_vs_baseline"] = grouped["M_sum"] - grouped["baseline_M_sum"]
    grouped["D_sum_vs_baseline"] = grouped["D_sum"] - grouped["baseline_D_sum"]
    grouped["delta_p_dispatch_sum_vs_baseline"] = grouped["delta_p_dispatch_sum"] - grouped["baseline_delta_p_dispatch_sum"]
    grouped["headroom_up_sum_vs_baseline"] = grouped["headroom_up_sum"] - grouped["baseline_headroom_up_sum"]
    return grouped


def _plot_area_totals(area_df: pd.DataFrame, *, stem: str) -> list[str]:
    if area_df.empty:
        return []

    set_thesis_style()
    formulation_ids = area_df["formulation_id"].dropna().astype(str).unique().tolist()
    area_names = area_df["area_name"].dropna().astype(str).unique().tolist()
    formulation_ids = sorted(formulation_ids, key=lambda value: (value == "", value))
    area_names = sorted(area_names)

    fig, axes = plt.subplots(1, 2, figsize=(11.0, 4.2))
    for ax, value_col, title, ylabel in [
        (axes[0], "M_sum", "Area-Wise Scheduled Inertia", "Area Sum of M [-]"),
        (axes[1], "D_sum", "Area-Wise Scheduled Damping", "Area Sum of D [-]"),
    ]:
        x = np.arange(len(area_names), dtype=float)
        width = 0.8 / max(len(formulation_ids), 1)
        for idx, formulation_id in enumerate(formulation_ids):
            subset = (
                area_df.loc[area_df["formulation_id"].astype(str) == formulation_id, ["area_name", value_col]]
                .set_index("area_name")
                .reindex(area_names)
            )
            values = pd.to_numeric(subset[value_col], errors="coerce").fillna(0.0).to_numpy()
            offset = (idx - (len(formulation_ids) - 1) / 2.0) * width
            ax.bar(x + offset, values, width=width, label=formulation_id, color=formulation_color(formulation_id), alpha=0.9)
        ax.set_xticks(x)
        ax.set_xticklabels(area_names)
        ax.set_title(title)
        ax.set_ylabel(ylabel)
    handles, labels = axes[0].get_legend_handles_labels()
    if handles:
        fig.legend(handles, labels, loc="upper center", ncol=min(len(labels), 5), frameon=False)
    fig.suptitle("Area-Wise VIS Allocation Comparison", y=1.02)
    paths = save_figure(fig, stem, FIGURES_DIR)
    plt.close(fig)
    return [str(path) for path in paths]


def _plot_unit_allocations(unit_df: pd.DataFrame, *, stem: str) -> list[str]:
    if unit_df.empty:
        return []

    set_thesis_style()
    focus = unit_df.loc[
        unit_df["formulation_id"].astype(str).isin(
            ["she_method_iii_dyn_fixed_md_with_reserve", "she_method_iv_vis_rted_full", "she_method_iv_vis_rted_area_tied"]
        )
    ].copy()
    if focus.empty:
        focus = unit_df.copy()

    methods = focus["formulation_id"].dropna().astype(str).unique().tolist()
    units = focus["regcv1_name"].dropna().astype(str).unique().tolist()
    units = sorted(units)

    fig, axes = plt.subplots(1, 2, figsize=(11.0, 4.2))
    for ax, value_col, title, ylabel in [
        (axes[0], "M_opt", "Per-Unit Scheduled Inertia", "M [-]"),
        (axes[1], "D_opt", "Per-Unit Scheduled Damping", "D [-]"),
    ]:
        x = np.arange(len(units), dtype=float)
        width = 0.8 / max(len(methods), 1)
        for idx, formulation_id in enumerate(methods):
            subset = (
                focus.loc[focus["formulation_id"].astype(str) == formulation_id, ["regcv1_name", value_col]]
                .set_index("regcv1_name")
                .reindex(units)
            )
            values = pd.to_numeric(subset[value_col], errors="coerce").fillna(0.0).to_numpy()
            offset = (idx - (len(methods) - 1) / 2.0) * width
            ax.bar(x + offset, values, width=width, label=formulation_id, color=formulation_color(formulation_id), alpha=0.9)
        ax.set_xticks(x)
        ax.set_xticklabels(units, rotation=20, ha="right")
        ax.set_title(title)
        ax.set_ylabel(ylabel)
        if value_col == "M_opt":
            ax.set_ylim(0.0, 8.0)
        elif value_col == "D_opt":
            ax.set_ylim(0.0, 6.0)
    handles, labels = axes[0].get_legend_handles_labels()
    if handles:
        fig.legend(handles, labels, loc="upper center", ncol=min(len(labels), 4), frameon=False)
    fig.suptitle("Per-Unit VIS Allocation: She-Style and Area-Tied Comparison", y=1.04)
    paths = save_figure(fig, stem, FIGURES_DIR)
    plt.close(fig)
    return [str(path) for path in paths]


def main() -> None:
    parser = argparse.ArgumentParser(description="Build area-wise VIS allocation tables and figures from a suite summary.")
    parser.add_argument(
        "--summary-json",
        default="results/thesis_optimization_results/results/area_vis_comparison_summary.json",
        help="Suite summary JSON written by scheduling.run_experiment_suite.",
    )
    parser.add_argument(
        "--baseline-id",
        default="she_method_i_rted",
        help="Baseline formulation id for delta columns.",
    )
    parser.add_argument(
        "--case",
        default="",
        help="Optional explicit system case path. If omitted, the first run's system_case is used.",
    )
    parser.add_argument(
        "--stem",
        default="area_vis_comparison",
        help="Output stem used for tables and figures.",
    )
    args = parser.parse_args()

    ensure_output_dirs()
    summary_path = Path(args.summary_json)
    if not summary_path.is_absolute():
        summary_path = (ROOT / summary_path).resolve()
    if not summary_path.exists():
        raise FileNotFoundError(f"Suite summary JSON not found: {summary_path}")

    ibr_df, system_case = _load_ibr_rows(summary_path)
    if ibr_df.empty:
        raise RuntimeError(f"No IBR dispatch rows found in suite summary: {summary_path}")

    raw_case = str(args.case).strip() or str(system_case or "").strip()
    if not raw_case:
        raise ValueError("Could not resolve system case path from --case or suite summary.")
    case_path = Path(raw_case)
    if not case_path.is_absolute():
        case_path = (ROOT / case_path).resolve()
    if not case_path.exists():
        raise FileNotFoundError(f"System case not found: {case_path}")

    area_map = _load_area_map(case_path)
    unit_df = ibr_df.merge(area_map, on="index", how="left")
    unit_df["area_name"] = unit_df["area_name"].fillna("UNKNOWN")
    unit_df["area_label"] = unit_df["area_name"].astype(str) + " / " + unit_df["regcv1_name"].astype(str)

    area_df = _area_summary(unit_df.copy(), baseline_id=str(args.baseline_id))

    stem = str(args.stem)
    outputs = {
        "unit_table": _write_table_bundle(f"{stem}_unit_allocations", unit_df),
        "area_table": _write_table_bundle(f"{stem}_by_area", area_df),
        "figures": [],
    }
    outputs["figures"].extend(_plot_area_totals(area_df, stem=f"{stem}_totals"))
    outputs["figures"].extend(_plot_unit_allocations(unit_df, stem=f"{stem}_units"))

    summary_out = TABLES_DIR / f"{stem}_summary.json"
    with summary_out.open("w", encoding="utf-8") as f:
        json.dump(outputs, f, indent=2)

    print(f"[area_vis_analysis] unit_rows={len(unit_df)} area_rows={len(area_df)}")
    print(f"[area_vis_analysis] summary={summary_out}")


if __name__ == "__main__":
    main()
